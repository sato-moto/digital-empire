"""
parse_excel.py — Excel スライドデータを JSON に変換

Usage:
    python parse_excel.py <input.xlsx> <output.json>
"""
import sys, json, io
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


def parse(excel_path: str) -> list[dict]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    slides = []

    # 新テンプレート: A=章, B=スライド番号, C=スライドタイトル, D=スライド用テキスト, E=ミニ台本
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row) + [None] * 6
        chapter = cells[0]   # A列: 章番号
        no = cells[1]        # B列: スライド番号
        title = cells[2]     # C列: スライドタイトル
        body = cells[3]      # D列: スライド用テキスト
        script = cells[4]    # E列: ミニ台本

        title = str(title or "").strip()
        body = str(body or "").strip()
        script = str(script or "").strip()

        if not title and not body:
            continue

        if title == "講座名":
            # 全体カバー: D列を改行分割 → 1行目=タイトル, 2行目=サブタイトル
            parts = [p.strip() for p in body.split("\n") if p.strip()]
            slides.append({
                "type": "cover",
                "title": parts[0] if parts else "",
                "subtitle": parts[1] if len(parts) > 1 else "",
                "body": "",
                "script": script,
            })
        elif title in ("タイトルスライド", "章タイトルスライド"):
            # 章タイトル: D列を改行分割 → 1行目=タイトル, 2行目=サブタイトル
            parts = [p.strip() for p in body.split("\n") if p.strip()]
            slides.append({
                "type": "chapter",
                "title": parts[0] if parts else "",
                "subtitle": parts[1] if len(parts) > 1 else "",
                "body": "",
                "script": script,
            })
        else:
            slides.append({
                "type": "content",
                "title": title,
                "body": body,
                "script": script,
            })

    wb.close()
    return slides


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python parse_excel.py <input.xlsx> <output.json>")
        sys.exit(1)

    data = parse(sys.argv[1])
    out = Path(sys.argv[2])
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps({"slides": data}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] {len(data)} slides -> {out}")
