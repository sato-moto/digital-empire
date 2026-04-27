# -*- coding: utf-8 -*-
"""course a teen の slide-design.xlsx に対してラウンド2修正

S13, S16, S18, S20, S21, S26 の Body を修正
"""
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent / "MATH-M-01-C-a/output_teen"
SD = ROOT / "slide-design.xlsx"

wb = load_workbook(SD)
ws = wb.active

# 各スライド番号 → 新Body
new_bodies = {
    # S13: プロンプト例を改行分割（各行<50字でheading扱いに）
    13: (
        "【プロンプト例】\n"
        "三角形の合同証明で、どの合同条件を使えばいいか迷う\n"
        "問題文から条件を絞り込むための\n"
        "最初の『考え方のステップ』を教えて"
    ),
    # S16: プロンプト例を改行分割（英語化を防ぐ）
    16: (
        "【プロンプト例】\n"
        "図形の問題で補助線をどこに引けばいいか分からない\n"
        "補助線を引くときの『よくあるパターン』や『狙い』を\n"
        "3つくらい挙げて"
    ),
    # S18: 内容追加で余白を埋める
    18: (
        "Geminiの画像生成について\n"
        "図解や補助線の画像を出せる\n"
        "ただし数学的に厳密ではない\n"
        "あくまでイメージの補助\n"
        "メインはテキスト壁打ち"
    ),
    # S20: プロンプト例を改行分割
    20: (
        "【プロンプト例】\n"
        "今から自分で書いた証明を送るから\n"
        "論理の飛躍がないか\n"
        "もっとシンプルに書ける部分がないか\n"
        "コーチとして厳しく添削して"
    ),
    # S21: 番号付き＋ヘッダー追加で文字ラベル化
    21: (
        "3つの典型ミス\n"
        "1. 明らかにで論拠を飛ばす\n"
        "2. 対応する頂点の順序が逆\n"
        "3. 仮定の使い忘れ"
    ),
    # S26: 「今のあなたへ:」「許可は手に入れた」削除
    26: (
        "あとは対話を始める番\n"
        "続きで待っている"
    ),
}

# 行のスライド番号で見つけて Body 列(D列=index 4)を更新
header_row = 1
for row in ws.iter_rows(min_row=2):
    no_cell = row[1]  # スライド番号
    body_cell = row[3]  # Body
    try:
        no = int(no_cell.value)
    except (TypeError, ValueError):
        continue
    if no in new_bodies:
        old = body_cell.value
        body_cell.value = new_bodies[no]
        print(f"[S{no}] 更新")
        print(f"  before: {(old or '')[:80].replace(chr(10), ' / ')}")
        print(f"  after:  {new_bodies[no][:80].replace(chr(10), ' / ')}")
        print()

wb.save(SD)
print(f"[OK] {SD}")
