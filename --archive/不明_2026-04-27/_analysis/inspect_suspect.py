# -*- coding: utf-8 -*-
"""疑わしいスライドの中身を実際に取り出して目視レベルで確認"""
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent

# 検査対象（KWマッチ低かったやつ）
SUSPECTS = {
    'a': [8, 13, 16, 21],
    'b': [13, 16, 18, 22, 23, 29],
}


def get_row(ws, slide_no):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[1]).strip() == str(slide_no).strip():
            return row
    return None


for course in ['a', 'b']:
    sd = load_workbook(ROOT / f"MATH-M-01-C-{course}/output_teen/slide-design.xlsx").active
    sc = load_workbook(ROOT / f"MATH-M-01-C-{course}/output_teen/script.xlsx").active

    print(f"\n{'#' * 70}")
    print(f"# Course {course}")
    print(f"{'#' * 70}")

    for n in SUSPECTS[course]:
        sd_row = get_row(sd, n)
        sc_row = get_row(sc, n)
        if not sd_row or not sc_row:
            continue
        ch, no, title, body, mini = sd_row[:5]
        _, _, talk = sc_row[:3]

        print(f"\n──── slide {no} (Ch{ch}) ────")
        print(f"【タイトル】{title}")
        print(f"\n【Body】\n{body}")
        print(f"\n【MiniScript（slide-design内）】\n{(mini or '')[:300]}")
        print(f"\n【script.xlsx 台本】\n{(talk or '')[:500]}")
        print(f"\n[整合判定材料]")
        print(f"  - Body に書いてある主旨: {(body or '')[:80].replace(chr(10), ' / ')}")
        print(f"  - 台本の冒頭: {(talk or '')[:80].replace(chr(10), ' / ')}")
        print()
