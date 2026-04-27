# -*- coding: utf-8 -*-
"""course a teen の slide-design.xlsx と script.xlsx に修正を適用する

修正一覧:
  (1) S1, S6, S11, S22: サブタイトル削除
  (2) Ch1(slides 2-5) と Ch2(slides 6-10) を入れ替え
  (3) S5(入替後はS10): 時間表記「10分後」削除
  (4) S12: Body冒頭「ポイント:」削除
  (5) S13, S20: 「【プロンプト例】」キャプション追加

teen版のみ・course aのみ。
"""
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent / "MATH-M-01-C-a/output_teen"
SD_PATH = ROOT / "slide-design.xlsx"
SC_PATH = ROOT / "script.xlsx"


def load_rows(path):
    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))
    return header, rows, wb


def write_rows(wb, header, rows, path):
    ws = wb.active
    # 既存行クリア
    ws.delete_rows(2, ws.max_row)
    # ヘッダはそのまま、データ書き込み
    for r in rows:
        ws.append(r)
    wb.save(path)


# ── slide-design ──
sd_header, sd_rows, sd_wb = load_rows(SD_PATH)
print(f"[INFO] slide-design 読込: {len(sd_rows)} 行")

# 元データ抜き出し（slide_no -> row）
def by_no(rows, no):
    for r in rows:
        if str(r[1]).strip() == str(no).strip():
            return r
    return None


# 構造化: スライド番号別
orig = {}
for r in sd_rows:
    orig[r[1]] = list(r)

# (1) サブタイトル削除（章=0/1の章タイトル系: S1=Cover, S6/S11/S22=章タイトル）
# Body の1行目のみ残す
def strip_subtitle(row):
    body = row[3] or ""
    first_line = body.split("\n", 1)[0]
    row[3] = first_line
    return row

# (2) Ch swap: 元S2-5(Ch1) と S6-10(Ch2) を入替
# 入替後: 新S2=旧S6, 新S3-6=旧S7-10, 新S7=旧S2, 新S8-10=旧S3-5

new_rows = []
# Cover (S1)
s1 = list(orig[1])
s1 = strip_subtitle(s1)
new_rows.append(s1)

# 新S2 = 旧S6 (元Ch2 章タイトル) — Ch=1 に書換、サブタイトル削除
s2 = list(orig[6])
s2[0] = 1
s2[1] = 2
s2 = strip_subtitle(s2)
new_rows.append(s2)

# 新S3-6 = 旧S7-10 (Ch2 body) — Ch=1 に書換
for i, old_no in enumerate([7, 8, 9, 10], start=3):
    r = list(orig[old_no])
    r[0] = 1
    r[1] = i
    new_rows.append(r)

# 新S7 = 旧S2 (元Ch1 章タイトル) — Ch=2 に書換、サブタイトル削除
s7 = list(orig[2])
s7[0] = 2
s7[1] = 7
s7 = strip_subtitle(s7)
new_rows.append(s7)

# 新S8-10 = 旧S3-5 (Ch1 body) — Ch=2 に書換
# (3) 新S10(=旧S5) の「10分後」表記を削除
for i, old_no in enumerate([3, 4, 5], start=8):
    r = list(orig[old_no])
    r[0] = 2
    r[1] = i
    if old_no == 5:  # 「10分後に手にする3つの型」
        # タイトル変更
        r[2] = "ここで手にする3つの型"
        # ミニ台本も修正
        if r[4]:
            r[4] = r[4].replace("この10分で、", "ここで、")
            r[4] = r[4].replace("この10分で", "ここで")
            r[4] = r[4].replace("10分で、", "")
            r[4] = r[4].replace("10分で", "")
    new_rows.append(r)

# 新S11 (=旧S11): Ch3 章タイトル, サブタイトル削除
s11 = list(orig[11])
s11 = strip_subtitle(s11)
new_rows.append(s11)

# 新S12 (=旧S12): 「ポイント:」削除
# (4) Body 冒頭の「ポイント:」 (改行含み) を削除
s12 = list(orig[12])
if s12[3]:
    body = s12[3]
    if body.startswith("ポイント:\n") or body.startswith("ポイント：\n"):
        body = body.split("\n", 1)[1]
    s12[3] = body
new_rows.append(s12)

# 新S13 (=旧S13): 「【プロンプト例】」追加
# (5) S13: Body 先頭にキャプション
s13 = list(orig[13])
if s13[3]:
    s13[3] = f"【プロンプト例】\n{s13[3]}"
new_rows.append(s13)

# S14-S19 (=旧S14-S19): そのまま
for n in [14, 15, 16, 17, 18, 19]:
    new_rows.append(list(orig[n]))

# 新S20 (=旧S20): 「【プロンプト例】」追加
s20 = list(orig[20])
if s20[3]:
    s20[3] = f"【プロンプト例】\n{s20[3]}"
new_rows.append(s20)

# S21 (=旧S21): そのまま
new_rows.append(list(orig[21]))

# S22 (=旧S22): Ch4 章タイトル, サブタイトル削除
s22 = list(orig[22])
s22 = strip_subtitle(s22)
new_rows.append(s22)

# S23-S26: そのまま
for n in [23, 24, 25, 26]:
    new_rows.append(list(orig[n]))

# 上書き保存
print(f"[INFO] 新スライド数: {len(new_rows)}")
write_rows(sd_wb, sd_header, new_rows, SD_PATH)
print(f"[OK] {SD_PATH}")


# ── script.xlsx ──
sc_header, sc_rows, sc_wb = load_rows(SC_PATH)
print(f"\n[INFO] script 読込: {len(sc_rows)} 行")
sc_orig = {}
for r in sc_rows:
    sc_orig[int(str(r[1]).strip())] = list(r)

new_sc = []
# 新S1 = 旧S1（Cover）
s1 = list(sc_orig[1])
# (3) 補足: cover script の「10分後」も削除
if s1[2]:
    s1[2] = s1[2].replace("10分後、", "受講後、")
    s1[2] = s1[2].replace("10分後", "受講後")
new_sc.append(s1)

# 新S2 = 旧S6 (Ch=1)
s = list(sc_orig[6]); s[0] = 1; s[1] = 2; new_sc.append(s)
# 新S3-6 = 旧S7-10 (Ch=1)
for i, old_no in enumerate([7, 8, 9, 10], start=3):
    s = list(sc_orig[old_no]); s[0] = 1; s[1] = i; new_sc.append(s)
# 新S7 = 旧S2 (Ch=2)
s = list(sc_orig[2]); s[0] = 2; s[1] = 7; new_sc.append(s)
# 新S8-9 = 旧S3-4 (Ch=2)
for i, old_no in enumerate([3, 4], start=8):
    s = list(sc_orig[old_no]); s[0] = 2; s[1] = i; new_sc.append(s)
# 新S10 = 旧S5 (Ch=2) — 時間表記削除
s = list(sc_orig[5])
s[0] = 2; s[1] = 10
if s[2]:
    s[2] = s[2].replace("この10分で、", "ここで、")
    s[2] = s[2].replace("この10分で", "ここで")
    s[2] = s[2].replace("10分で、", "")
    s[2] = s[2].replace("10分で", "")
    s[2] = s[2].replace("10分後、", "受講後、")
    s[2] = s[2].replace("10分後", "受講後")
new_sc.append(s)

# 新S11-S26 = 旧S11-S26（番号維持）
for n in range(11, 27):
    new_sc.append(list(sc_orig[n]))

print(f"[INFO] 新台本数: {len(new_sc)}")
write_rows(sc_wb, sc_header, new_sc, SC_PATH)
print(f"[OK] {SC_PATH}")

print("\n[DONE] modify_course_a_teen 完了")
