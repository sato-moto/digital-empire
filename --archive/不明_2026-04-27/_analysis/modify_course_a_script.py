# -*- coding: utf-8 -*-
"""course a teen の script.xlsx のみを修正する

slide-design は既に modify_course_a_teen.py で適用済み。
script.xlsx に対して、章入れ替え＋時間表記削除を反映。
"""
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent / "MATH-M-01-C-a/output_teen"
SC_PATH = ROOT / "script.xlsx"

wb = load_workbook(SC_PATH)
ws = wb.active
header = [c.value for c in ws[1]]

sc_orig = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    no = int(str(row[1]).strip())
    sc_orig[no] = list(row)

print(f"[INFO] script 読込: {len(sc_orig)} 行")

new_sc = []

# 新S1 = 旧S1 (Cover) — 「10分後」削除
s = list(sc_orig[1])
if s[2]:
    s[2] = s[2].replace("10分後、", "受講後、").replace("10分後", "受講後")
new_sc.append(s)

# 新S2 = 旧S6 (Ch=1 章タイトル → 章数字を1に変更)
s = list(sc_orig[6]); s[0] = 1; s[1] = 2
new_sc.append(s)

# 新S3-6 = 旧S7-10 (Ch=1 body)
for i, old_no in enumerate([7, 8, 9, 10], start=3):
    s = list(sc_orig[old_no]); s[0] = 1; s[1] = i
    new_sc.append(s)

# 新S7 = 旧S2 (Ch=2 章タイトル)
s = list(sc_orig[2]); s[0] = 2; s[1] = 7
new_sc.append(s)

# 新S8-9 = 旧S3-4 (Ch=2 body 1-2)
for i, old_no in enumerate([3, 4], start=8):
    s = list(sc_orig[old_no]); s[0] = 2; s[1] = i
    new_sc.append(s)

# 新S10 = 旧S5 (Ch=2 body 3) — 時間表記削除
s = list(sc_orig[5])
s[0] = 2; s[1] = 10
if s[2]:
    s[2] = s[2].replace("この10分で、", "ここで、").replace("この10分で", "ここで")
    s[2] = s[2].replace("10分で、", "").replace("10分で", "")
    s[2] = s[2].replace("10分後、", "受講後、").replace("10分後", "受講後")
new_sc.append(s)

# 新S11-S26 = 旧S11-S26（番号維持）
for n in range(11, 27):
    new_sc.append(list(sc_orig[n]))

# 上書き
ws.delete_rows(2, ws.max_row)
for r in new_sc:
    ws.append(r)
wb.save(SC_PATH)
print(f"[OK] {SC_PATH} ({len(new_sc)} 行)")
print()
print("[新台本]")
for r in new_sc:
    talk = (r[2] or "")[:50].replace("\n", " ")
    print(f"  ch={r[0]} no={r[1]:>2} talk={talk}")
