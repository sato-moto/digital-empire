# -*- coding: utf-8 -*-
"""slide-design.xlsx vs script.xlsx の整合性チェック

3軸で評価:
  1. 行数・スライド番号の一致
  2. キーワード重複度（slide-design の Body/Title と script の本文に共通の核心用語があるか）
  3. 文量バランス（章タイトル・カバーは台本やや軽め、本文スライドは適切な厚み）
"""
import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent

# 図形・数学・本SET特有のキーワード（核心用語）— これがslide↔scriptの両方にあれば整合
CORE_KEYWORDS = [
    # 一般的な型
    "方針", "補助線", "添削", "壁打ち", "プロンプト", "型", "AI", "コーチ",
    # 図形分野
    "合同", "証明", "条件", "仮定", "対頂角", "錯角", "共通", "平行",
    "SAS", "SSS", "ASA", "三角形", "四角形", "対角線", "円周角", "相似",
    # 学習姿勢
    "思考", "丸暗記", "自走", "1歩", "次の一歩", "整理", "材料",
    # 講座構造
    "Step", "ステップ", "図", "ヒント", "チェックリスト",
    # NotebookLM接続
    "NotebookLM", "ストック", "ログ", "ノート",
    # 失敗パターン
    "丸投げ", "見た目", "ひらめき",
    # ツール
    "Gemini", "ChatGPT",
    # その他
    "実戦", "対話", "記号",
]


def extract_keywords(text):
    if not text:
        return set()
    return {kw for kw in CORE_KEYWORDS if kw in text}


def analyze(course):
    print(f"\n{'=' * 60}")
    print(f"  Course {course}: slide-design ↔ script 整合性チェック")
    print(f"{'=' * 60}")

    sd = load_workbook(ROOT / f"MATH-M-01-C-{course}/output_teen/slide-design.xlsx").active
    sc = load_workbook(ROOT / f"MATH-M-01-C-{course}/output_teen/script.xlsx").active

    sd_rows = list(sd.iter_rows(min_row=2, values_only=True))
    sc_rows = list(sc.iter_rows(min_row=2, values_only=True))

    issues = []

    # 1. 行数一致
    if len(sd_rows) != len(sc_rows):
        issues.append(f"❌ 行数不一致: slide-design={len(sd_rows)}, script={len(sc_rows)}")

    # 各スライドごとのチェック
    print(f"\n【スライドごとの整合性】")
    print(f"{'#':>3} {'章':>2} {'タイトル':<30} {'共有KW':>5} {'KWカバー率':>10} {'評価'}")
    print('-' * 100)

    for i, (sd_r, sc_r) in enumerate(zip(sd_rows, sc_rows)):
        ch_sd, no_sd, title, body, mini = sd_r[:5]
        ch_sc, no_sc, talk = sc_r[:3]

        # スライド番号一致（int/str両対応）
        if str(no_sd).strip() != str(no_sc).strip():
            issues.append(f"❌ s{i + 1}: スライド番号不一致 sd={no_sd!r} sc={no_sc!r}")
            continue

        is_chapter = (title == "章タイトルスライド")
        is_cover = (title == "講座名")

        title_kw = extract_keywords(title or "")
        body_kw = extract_keywords(body or "")
        talk_kw = extract_keywords(talk or "")

        slide_kw = title_kw | body_kw  # スライド側のキーワード
        shared = slide_kw & talk_kw
        coverage = len(shared) / len(slide_kw) if slide_kw else 0

        title_short = (title or "")[:28].replace("\n", " ")

        # 評価ロジック
        if is_cover:
            label = "[COVER]"
        elif is_chapter:
            label = "[CHAPTER]"
        elif coverage >= 0.7:
            label = "✅ OK"
        elif coverage >= 0.4:
            label = "⚠️ 弱い"
            issues.append(f"⚠️ s{no_sd} '{title_short}': KW共有 {len(shared)}/{len(slide_kw)} ({coverage*100:.0f}%)")
        else:
            label = "❌ 不整合"
            issues.append(f"❌ s{no_sd} '{title_short}': KW共有 {len(shared)}/{len(slide_kw)} ({coverage*100:.0f}%)\n    slide_kw={slide_kw}\n    talk_kw={talk_kw}")

        print(f"{no_sd:>3} {ch_sd:>2} {title_short:<30} {len(shared):>5} {coverage*100:>8.0f}%  {label}")

        # 文量チェック（台本）
        talk_len = len(talk or "")
        if not is_cover and not is_chapter and talk_len < 150:
            issues.append(f"⚠️ s{no_sd}: 台本短い（{talk_len}字） -- 本文スライドは200字以上目安")

    # サマリー
    print(f"\n【サマリー】")
    if not issues:
        print("  ✅ 大きな問題なし")
    else:
        print(f"  検出: {len(issues)} 件")
        for x in issues[:30]:
            print(f"    {x}")


if __name__ == "__main__":
    analyze('a')
    analyze('b')
