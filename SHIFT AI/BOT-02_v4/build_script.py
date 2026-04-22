# -*- coding: utf-8 -*-
"""
slide-design.xlsx の E列（ミニ台本）を抽出して、
script.csv / script.xlsx として台本ファイルを生成する。
"""
import csv
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

BASE = Path(__file__).parent
SRC = BASE / "slide-design.xlsx"
CSV_PATH = BASE / "script.csv"
XLSX_PATH = BASE / "script.xlsx"


def extract():
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        chapter, no, title, body, script = (row + (None,) * 5)[:5]
        if not title and not body:
            continue
        rows.append((chapter, no, str(script or "").strip()))
    wb.close()
    return rows


def write_csv(path: Path, rows):
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        w.writerow(["章", "スライド番号", "台本"])
        for r in rows:
            w.writerow(r)
    print(f"[OK] CSV: {path}")


def write_xlsx(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "script"
    ws.append(["章", "スライド番号", "台本"])
    for r in rows:
        ws.append(list(r))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 120
    wrap = Alignment(wrap_text=True, vertical="top")
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
        for c in r:
            c.alignment = wrap
    wb.save(path)
    print(f"[OK] XLSX: {path}")


if __name__ == "__main__":
    rows = extract()
    write_csv(CSV_PATH, rows)
    write_xlsx(XLSX_PATH, rows)
    print(f"Total rows: {len(rows)}")
