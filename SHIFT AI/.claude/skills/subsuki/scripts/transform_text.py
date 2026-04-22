# -*- coding: utf-8 -*-
"""
transform_text.py — サブスキちゃん用 テキスト変換ツール

役割:
  - slide-design.xlsx の D列（Body）を辞書マッチで学校生活寄りに置換
  - script.xlsx の C列（台本）を同辞書で置換
  - thumbnail-prompt.txt を辞書＋追加指示注入で学校生活寄りに変換
  - CSV/XLSX は output_teen/ 配下に新規出力（元ファイルは上書きしない）
  - 変換箇所は diff_report.json に Before/After で記録

辞書マッチ方針:
  - dictionaries/business_to_school.json を参照
  - 最長マッチ優先、大文字小文字を区別しない
  - 辞書にない語彙は推測変換しない（そのまま維持）

Usage:
  python transform_text.py \\
      --slide-design <path> \\
      --script <path> \\
      --thumbnail-prompt <path> \\
      --output-dir <dir>
"""
import argparse
import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

# ─── 定数 ──────────────────────────────────────────────
SKILL_DIR = Path(__file__).resolve().parent.parent
DICT_PATH = SKILL_DIR / "dictionaries" / "business_to_school.json"

THUMBNAIL_TEEN_PREAMBLE = """
# 中高生向けテイスト指示（サブスキちゃん後付け）
このサムネイル画像は中高生向けオンライン学習教材の広告として使用される。
以下を必ず守ること：
- 登場する人物は制服またはカジュアル服の学生（中学生・高校生・大学生）のみ。スーツ・オフィスカジュアル・ビジネスパーソンは一切描かない
- 背景・小物は学校・教室・図書室・自宅の机・リビング学習・スマホ学習など**学校生活を想起させる空間**に限定する。オフィス・会議室・商談シーンは描かない
- 雰囲気は中高生が「自分ごと」と感じられる親しみやすさを最優先にする
- キャラクターは13〜18歳のリアルな等身感。子供っぽいデフォルメ（2〜3頭身）は避ける
"""


# ─── ユーティリティ ────────────────────────────────────
def load_dict(path: Path):
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    mappings = data.get("mappings", [])
    # 最長マッチ優先
    mappings.sort(key=lambda m: len(m["from"]), reverse=True)
    return mappings


def apply_dict(text: str, mappings, slide_no=None, field=None, diffs=None):
    """辞書マッチで text 内の語彙を置換。変更箇所は diffs に追記"""
    if not text:
        return text
    new_text = text
    for m in mappings:
        frm = m["from"]
        to = m["to"]
        # 大文字小文字区別なし → 正規表現で
        pattern = re.compile(re.escape(frm), re.IGNORECASE)
        matches = list(pattern.finditer(new_text))
        if matches:
            if diffs is not None:
                for match in matches:
                    diffs.append({
                        "slide": slide_no,
                        "field": field,
                        "before": match.group(0),
                        "after": to,
                        "context": m.get("context", ""),
                    })
            new_text = pattern.sub(to, new_text)
    return new_text


def format_body_lines(text: str, slide_no=None, diffs=None):
    """Body の各行を extractDiagramContent が拾える形に整形する。

    ルール（D-3案）:
      - 各行頭に「・」を付ける（既に「・」「-」「数字.」「数字)」「数字）」で始まる行はそのまま）
      - 各行末尾の句点「。」を削除する
      - 空行はスキップ

    目的:
      slide-generator の extractDiagramContent() は行頭「・/-/数字.」または
      50文字未満＋句点なしの行だけを抽出する。Body が句点で終わる文章だと
      bullets にも headings にも入らず、Gemini に「テキストなし」と伝わる。
      全行を「・」始まり＋句点なしに統一することで、抽出が確実になる。
    """
    if not text:
        return text
    lines = text.split("\n")
    formatted = []
    changed = False
    for original in lines:
        line = original.rstrip()
        if not line.strip():
            formatted.append("")
            continue
        # 末尾の句点を削除（連続「。。」にも対応）
        while line.endswith("。"):
            line = line[:-1]
        # 行頭が「・」「-」「数字.」「数字)」「数字）」で始まる場合はそのまま
        if re.match(r"^[・\-]|^\d+[.\)）]", line):
            new_line = line
        else:
            new_line = "・" + line
        if new_line != original:
            changed = True
        formatted.append(new_line)
    if changed and diffs is not None:
        diffs.append({
            "slide": slide_no,
            "field": "BodyFormat",
            "before": "（従前の文章形式）",
            "after": "（行頭「・」＋句点削除に整形）",
            "context": "D-3案統一ルール適用",
        })
    return "\n".join(formatted)


# ─── XLSX 変換 ─────────────────────────────────────────
def transform_slide_design(src: Path, dst: Path, mappings, diffs):
    """slide-design.xlsx の D列(Body)を辞書変換し、dst へ保存"""
    wb = load_workbook(src)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        # 章, No, Title, Body, Script
        chapter_cell, no_cell, title_cell, body_cell, script_cell = row[:5]
        slide_no = no_cell.value
        if body_cell.value:
            # Step 1: 辞書マッチで語彙置換
            body_text = apply_dict(
                str(body_cell.value), mappings,
                slide_no=slide_no, field="Body", diffs=diffs,
            )
            # Step 2: D-3案 Body 整形（行頭「・」＋句点削除）
            body_text = format_body_lines(body_text, slide_no=slide_no, diffs=diffs)
            body_cell.value = body_text
        if script_cell.value:
            script_cell.value = apply_dict(
                str(script_cell.value), mappings,
                slide_no=slide_no, field="MiniScript", diffs=diffs,
            )

    # 書式維持のためのwrap
    wrap = Alignment(wrap_text=True, vertical="top")
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=5):
        for c in r:
            c.alignment = wrap

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    wb.close()
    return dst


def transform_script(src: Path, dst: Path, mappings, diffs):
    """script.xlsx の C列(台本)を辞書変換し、dst へ保存"""
    wb = load_workbook(src)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        # 章, スライド番号, 台本
        chapter_cell, no_cell, script_cell = row[:3]
        slide_no = no_cell.value
        if script_cell.value:
            script_cell.value = apply_dict(
                str(script_cell.value), mappings,
                slide_no=slide_no, field="Script", diffs=diffs,
            )

    wrap = Alignment(wrap_text=True, vertical="top")
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
        for c in r:
            c.alignment = wrap

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    wb.close()
    return dst


# ─── CSV 連動更新 ──────────────────────────────────────
def sync_slide_design_csv(src_xlsx: Path, dst_csv: Path):
    """xlsx の内容を CSV として書き出し（CSVとxlsxのペア同期ルール遵守）"""
    wb = load_workbook(src_xlsx, data_only=True)
    ws = wb.active
    dst_csv.parent.mkdir(parents=True, exist_ok=True)
    with dst_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        for row in ws.iter_rows(values_only=True):
            w.writerow(list(row))
    wb.close()


def sync_script_csv(src_xlsx: Path, dst_csv: Path):
    wb = load_workbook(src_xlsx, data_only=True)
    ws = wb.active
    dst_csv.parent.mkdir(parents=True, exist_ok=True)
    with dst_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        for row in ws.iter_rows(values_only=True):
            w.writerow(list(row))
    wb.close()


# ─── サムネプロンプト変換 ──────────────────────────────
def transform_thumbnail_prompt(src: Path, dst: Path, mappings, diffs):
    """thumbnail-prompt.txt を辞書マッチ＋追加指示注入で変換"""
    content = src.read_text(encoding="utf-8")
    transformed = apply_dict(content, mappings, slide_no="thumbnail", field="ThumbPrompt", diffs=diffs)
    # 末尾に中高生向け追加指示を注入
    transformed += "\n\n" + THUMBNAIL_TEEN_PREAMBLE.strip() + "\n"
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_text(transformed, encoding="utf-8")


# ─── メイン ─────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="サブスキちゃん テキスト変換ツール")
    ap.add_argument("--slide-design", required=True, help="元の slide-design.xlsx のパス")
    ap.add_argument("--script", required=True, help="元の script.xlsx のパス")
    ap.add_argument("--thumbnail-prompt", required=False, help="元の thumbnail-prompt.txt のパス（任意）")
    ap.add_argument("--output-dir", required=True, help="出力先ディレクトリ（例: output_teen/）")
    args = ap.parse_args()

    src_slide = Path(args.slide_design)
    src_script = Path(args.script)
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    mappings = load_dict(DICT_PATH)
    diffs = []

    # 1) slide-design.xlsx 変換
    dst_slide = out_dir / "slide-design.xlsx"
    transform_slide_design(src_slide, dst_slide, mappings, diffs)
    sync_slide_design_csv(dst_slide, out_dir / "slide-design.csv")
    print(f"[OK] slide-design.xlsx -> {dst_slide}")

    # 2) script.xlsx 変換
    dst_script = out_dir / "script.xlsx"
    transform_script(src_script, dst_script, mappings, diffs)
    sync_script_csv(dst_script, out_dir / "script.csv")
    print(f"[OK] script.xlsx -> {dst_script}")

    # 3) thumbnail-prompt.txt 変換（任意）
    if args.thumbnail_prompt:
        src_prompt = Path(args.thumbnail_prompt)
        if src_prompt.exists():
            dst_prompt = out_dir / "thumbnail-prompt-teen.txt"
            transform_thumbnail_prompt(src_prompt, dst_prompt, mappings, diffs)
            print(f"[OK] thumbnail-prompt.txt -> {dst_prompt}")
        else:
            print(f"[SKIP] thumbnail-prompt not found: {src_prompt}")

    # 4) diff_report.json 保存
    diff_path = out_dir / "diff_report.json"
    with diff_path.open("w", encoding="utf-8") as f:
        json.dump({"diffs": diffs, "total": len(diffs)}, f, ensure_ascii=False, indent=2)
    print(f"[OK] diff_report.json -> {diff_path} ({len(diffs)} changes)")


if __name__ == "__main__":
    main()
